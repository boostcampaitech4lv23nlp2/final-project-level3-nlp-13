from selenium import webdriver
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import time

import openpyxl
from openpyxl.styles import PatternFill, Color
from openpyxl import Workbook
from random import *
from selenium.webdriver.common.by import By
import os

# 정렬 방식 선택
# 1: 추천순
# 2: 최신순
# 기타: 정확도 순
def sort_kind(index):
    # 추천
    if index == 1:
        return 'vcount'
    # 최신순
    elif index == 2:
        return 'date'
    # 정확도
    else:
        return 'none'

keyword = '정국이'
driver.get("https://kin.naver.com/search/list.naver?query=%EC%A0%95%EA%B5%AD") 
#'https://kin.naver.com/search/list.naver?query=' + get_keyword(keyword))
time.sleep(uniform(0.1, 1.0))

page_index = 1
# 크롤링 시작 일자
f = '2023.01.24'
# 크롤링 종료 일자
t = '2023.01.26'
period_txt = "&period=" + f + ".%7C" + t + "."
count = 0

_sort_kind = sort_kind(2)
date = str(datetime.now()).replace('.', '_')
date = date.replace(' ', '_')

# URL 저장
file_name = keyword+ ".txt"
f = open(file_name, 'w')

page_url = []
while True:
    time.sleep(uniform(0.01, 1.0))
    driver.get('https://kin.naver.com/search/list.nhn?' + "&sort=" + _sort_kind + '&query=' + keyword + period_txt + "&section=kin" + "&page=" + str(page_index))
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')

    tags = soup.find_all('a', class_="_nclicks:kin.txt _searchListTitleAnchor")
    for tag in tags:
        url = str(tag).split(' ')[3]
        url = url.replace('href=', "")
        url = url.replace('"', "")
        url = url.replace('amp;', '')
        page_url.append(url)
        f.write(url + "\n")

    post_number = driver.find_element(By.CLASS_NAME,'number').text
    post_number = str(post_number).replace("(", "")
    post_number = str(post_number).replace(")", "")
    
    current_number = post_number.split('/')[0].split('-')[1]
    current_number = current_number.replace(',', '')
    total_number = post_number.split('/')[1]
    total_number = total_number.replace(',', '')

    if int(current_number) == int(total_number):
        break
    else:
        page_index += 1

filename = keyword + "_crawling_result.xlsx"

wb = Workbook()
sheet = wb.active
sheet.append(['제목', '질문', '답변'])

for j in range(1, 4):
    sheet.cell(row=1, column=j).fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')

for i in page_url:
    driver.get(i)
    title = driver.find_element(By.CLASS_NAME,'title').text
    try:
        question_txt = driver.find_element(By.CLASS_NAME,'c-heading__content').text
        
    except:
        question_txt = ""

    # 답변 리스트
    answer_list = driver.find_elements(By.CLASS_NAME,"se-main-container")
    
    for n, answer in enumerate(answer_list):
        texts = answer.find_elements(By.TAG_NAME, 'span')
        t = ""
        for i in texts:
            t += i.text

        if n == 0:
            sheet.append([title, question_txt, t])
        else:
            sheet.append(["", "", t])
    
    count += 1
    print(count)
    wb.save(filename)